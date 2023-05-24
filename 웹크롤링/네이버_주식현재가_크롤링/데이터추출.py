import requests
from bs4 import BeautifulSoup
import openpyxl

fpath = r'C:\Users\dhqhf\vs-workspace\웹크롤링\네이버_주식현재가_크롤링\data.xlsx'

wb = openpyxl.Workbook() #엑셀 파일 만들기

#컬럼 추가
ws = wb.active
ws['A1'] = '종목'
ws['B1'] = '현재가'
ws['C1'] = '평균매입가'
ws['D1'] = '잔고수량'
ws['E1'] = '평가금액'
ws['F1'] = '평가손익'
ws['G1'] = '수익률'

#현재가
#종목 코드 리스트
current_price_codes = [
    '005930', #삼성전자
    '000660', #SK하이닉스
    '035720' #카카오
]

row = 2
column_index = 2 
for code in current_price_codes:
    url = f"https://finance.naver.com/item/sise.naver?code={code}"
    response = requests.get(url) #url 사이트로 get 요청이 가고 그걸 response 변수에 받음
    html = response.text
    soup = BeautifulSoup(html, 'html.parser') #html 을 html.parser 로 번역한다
    price = soup.select_one("#_nowVal").text #태그 정보 하나만 가져올꺼면 select_one(), 여러개는 select(), "css선택자" 를 인자로 받음, .text 로 태그 안에 있는 문자열만 가져옴
    price = price.replace(',', '') #'97,700' 문자열을 숫자로 사용하기 위해 ',' 를 지워준다
    print(price)
    # ws[f'B{row}'] = int(price)
    #더 찾아보자
    column_letter = openpyxl.utils.get_column_letter(column_index)
    ws[f'{column_letter}{row}'].value = int(price)

    row = row + 1

wb.save(fpath) #저장
