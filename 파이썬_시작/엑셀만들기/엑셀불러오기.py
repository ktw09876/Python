import openpyxl

fpath = r'C:\Users\dhqhf\vs-workspace\엑셀만들기\참가자_data.xlsx'
wb = openpyxl.load_workbook(fpath) #엑셀 불러오기

#데이터 수정
ws = wb['오징어게임'] #시트선택
ws['A3'] = 456
ws['B3'] = '성기훈' 

wb.save(fpath)