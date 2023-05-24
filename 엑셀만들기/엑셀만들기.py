import openpyxl

wb = openpyxl.Workbook() #엑셀 파일 만들기

#데이터 추가
# ws = wb.create_sheet('오징어게임')
ws = wb.active
ws['A1'] = '참가번호'
ws['B1'] = '성명'

ws['A2'] = 1
ws['B2'] = '오일남'

wb.save(r'C:\Users\dhqhf\vs-workspace\엑셀만들기\참가자_data.xlsx') #저장